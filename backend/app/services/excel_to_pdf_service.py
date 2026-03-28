import os
from typing import List, Optional
from fastapi import UploadFile, HTTPException
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from app.utils.file_handler import save_upload_files, create_temp_file


async def run(files: List[UploadFile], options: Optional[dict] = None) -> str:
    paths: List[str] = []
    try:
        paths = await save_upload_files(files, allowed_exts=[".xlsx"])
        out_path, out_id = create_temp_file(".pdf")
        wb = load_workbook(paths[0], data_only=True)
        c = canvas.Canvas(out_path, pagesize=A4)
        width, height = A4
        x_margin = 0.75 * inch
        y_margin = 0.75 * inch
        y = height - y_margin
        for ws in wb.worksheets:
            y = height - y_margin
            for row in ws.iter_rows(values_only=True):
                line = " | ".join([str(v) if v is not None else "" for v in row]).strip()
                if y < y_margin + 12:
                    c.showPage()
                    y = height - y_margin
                c.drawString(x_margin, y, line)
                y -= 14
            c.showPage()
        c.save()
        return out_id
    finally:
        for p in paths:
            try:
                os.remove(p)
            except Exception:
                pass
